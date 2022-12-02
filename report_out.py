import csv
import re
import matplotlib.pyplot as plt
from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
import pandas as pd


class Tools:
    """Класс хранит в себе необходимые 'инструменты' для использования на протяжении всей программы,
    такие как: словарь и программа очистки текста.
    """
    rus_names = {'Название': 'name',
                 'Описание': 'description',
                 'Навыки': 'key_skills',
                 'Опыт работы': 'experience_id',
                 'Премиум-вакансия': 'premium',
                 'Компания': 'employer_name',
                 'Оклад': 'salary_from',
                 'Верхняя граница вилки оклада': 'salary_to',
                 'Оклад указан до вычета налогов': 'salary_gross',
                 'Идентификатор валюты оклада': 'salary_currency',
                 'Название региона': 'area_name',
                 'Дата публикации вакансии': 'published_at'}

    @staticmethod
    def prepare(text):
        """Очищает входную строку и возвращает очищенную для дальнейшего использования

            Args:
                text (str): Строка, которую нужно очистить

            Returns:
                str: Очищенная строка
        """
        text = re.sub(r"<[^>]+>", '', text)
        text = "; ".join(text.split('\n'))
        text = ' '.join(text.split())
        return text


class Vacancy:
    """Класс устанавливает все основные поля вакансии

        Attributes:
            name (str): Название вакансии
            salary (Salary): Комбинированная информация о зарплате
            area_name (str): Название региона
            published_at (str): Дата публикации вакансии
    """
    def __init__(self, name, salary, area_name, published_at):
        """Инициализирует объект Vacancy

            Args:
            name (str): Название вакансии
            salary (Salary): Комбинированная информация о зарплате
            area_name (str): Название региона
            published_at (str): Дата публикации вакансии
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    """Класс для представления зарплаты, также здесь находится словарь для перевода курсов - currency

        Attributes:
            salary_from (str): Нижняя граница вилки оклада
            salary_to (str): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
            salary_to_rub (int): Средняя зарплата в рублях
    """
    currency = {"AZN": 35.68,
                "BYR": 23.91,
                "EUR": 59.90,
                "GEL": 21.74,
                "KGS": 0.76,
                "KZT": 0.13,
                "RUR": 1,
                "UAH": 1.64,
                "USD": 60.66,
                "UZS": 0.0055}

    def __init__(self, salary_from, salary_to, salary_currency):
        """Инициализирует объект Salary

            Args:
                salary_from (str or int or float): Нижняя граница вилки оклада
                salary_to (str or int or float): Верхняя граница вилки оклада
                salary_currency (str): Валюта оклада
                salary_to_rub (int): Средняя зарплата в рублях
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_to_rub = Salary.currency_to_rub(salary_from, salary_to, salary_currency)

    @staticmethod
    def currency_to_rub(salary_from, salary_to, salary_currency):
        """Вычисляет среднюю зарплату из вилки и переводит в рубли, при помощи словаря - currency

            Args:
                salary_from (str): Нижняя вилка оклада
                salary_to (str): Верхняя вилка оклада
                salary_currency (str): Валюта оклада

            Returns:
                float: Средняя зарплата в рублях
        """
        return (float(salary_from) + float(salary_to)) / 2 * Salary.currency[salary_currency]


class DataSet:
    """Класс отвечает за чтение и подготовку данных из CSV-файла

        Attributes:
            file_name (str): Название файла
            vacancies_objects (list): Список вакансий
    """
    def __init__(self, file_name):
        """Инициализирует объект DataSet.
            Args:
                file_name (str): Название файла
                vacancies_objects (list): Список вакансий
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.prepare_data(file_name)

    @staticmethod
    def clear_csv(str_value):
        return " ".join(re.sub(r"<[^>]*>", "", str_value).split())

    @staticmethod
    def reader_csv(file_name):
        """Считывает csv файл.

            Args:
                file_name (str): Название файла

            Returns:
                tuple: Кортеж, состоящий из названий колонок и всех вакансий
        """
        file_csv = open(file_name, encoding="utf-8-sig")
        reader = csv.reader(file_csv)
        list_data = [x for x in reader]
        try:
            columns = list_data[0]
            return columns, list_data[1:]
        except:
            print("Пустой файл")
            exit()

    @staticmethod
    def prepare_data(file_name):
        """Отбирает вакансии без пустых ячеек и составляет лист вакансий

            Args:
                file_name (str): Название файла

            Returns:
                list: Лист, состоящий из вакансий
        """
        heads, data = DataSet.reader_csv(file_name)
        processed = [x for x in data if len(x) == len(heads) and "" not in x]
        vacancy_data = []
        for line in processed:
            dic = {}
            salary = []
            for i, item in enumerate(line):
                item = Tools.prepare(item)
                dic[heads[i]] = item
                if heads[i] in ['salary_from', 'salary_to', 'salary_currency']:
                    salary.append(Tools.prepare(item))
                    if len(salary) == 3:
                        dic["salary"] = Salary(salary[0], salary[1], salary[2])
            vacancy_data.append(Vacancy(dic["name"],
                                        dic["salary"],
                                        dic["area_name"],
                                        dic["published_at"]))
        return vacancy_data


class InputParam:
    """Класс отвечает за обработку параметров вводимых пользователем, а также за печать статистики

        Attributes:
            file_name (str): Название файла
            filter_param (str): Название профессии
    """
    def __init__(self):
        """Инициализирует объект InputConect

            Args:
                param (str): Параметры, введенные пользователем
        """
        self.params = InputParam.get_params()

    @staticmethod
    def get_params():
        """Возвращает заданные параметры, введенные пользователем

            Args:
                file_name (str): Название файла
                filter_param (str): Название профессии

            Returns:
                Tuple (str, str): Название файла и Название профессии
        """
        file_name = input("Введите название файла: ")
        vacancy = input("Введите название профессии: ")
        return file_name, vacancy

    @staticmethod
    def print_data(dictionary, key):
        """Печатает статистику и вызывает методы для формирования графиков и отчетов

            Args:
                dictionary (list): Список вакансий
                key (str): Название профессии
        """
        years = []
        for vacancy in dictionary:
            years.append(int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')))
        years.sort()
        years = list(range(min(years), max(years) + 1))

        salary_filter = {year: [] for year in years}
        vac_filter = {year: 0 for year in years}
        vac_sal_filter = {year: [] for year in years}
        vac_count_filter = {year: 0 for year in years}

        for vacancy in dictionary:
            year = int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
            salary_filter[year].append(vacancy.salary.salary_to_rub)
            vac_filter[year] += 1
            if key in vacancy.name:
                vac_sal_filter[year].append(vacancy.salary.salary_to_rub)
                vac_count_filter[year] += 1

        salary_filter = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                         salary_filter.items()}
        vac_sal_filter = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                          vac_sal_filter.items()}

        dic = {}
        for vacancy in dictionary:
            if vacancy.area_name in dic:
                dic[vacancy.area_name].append(vacancy.salary.salary_to_rub)
            else:
                dic[vacancy.area_name] = [vacancy.salary.salary_to_rub]

        area_filter = dic.items()
        area_filter = [x for x in area_filter if len(x[1]) / len(dictionary) > 0.01]
        area_filter.sort(key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        salary_cities_filter = {item[0]: int(sum(item[1]) / len(item[1])) for item in
                                area_filter[0: min(len(area_filter), 10)]}

        vacs_dic = {}
        for vacancy in dictionary:
            if vacancy.area_name in vacs_dic:
                vacs_dic[vacancy.area_name] += 1
            else:
                vacs_dic[vacancy.area_name] = 1

        count = {x: round(y / len(dictionary), 4) for x, y in vacs_dic.items()}
        count = {x: val for x, val in count.items() if val >= 0.01}
        vacs_cities = dict(sorted(count.items(), key=lambda item: item[1], reverse=True))
        others = sum(dict(list(vacs_cities.items())[11:]).values())
        vacs_cities = dict(list(vacs_cities.items())[:10])

        print('Динамика уровня зарплат по годам:', salary_filter)
        print('Динамика количества вакансий по годам:', vac_filter)
        print('Динамика уровня зарплат по годам для выбранной профессии:', vac_sal_filter)
        print('Динамика количества вакансий по годам для выбранной профессии:', vac_count_filter)
        print('Уровень зарплат по городам (в порядке убывания):', salary_cities_filter)
        print('Доля вакансий по городам (в порядке убывания):', vacs_cities)

        report = Report(salary_filter, vac_filter, vac_sal_filter, vac_count_filter, salary_cities_filter,
                        vacs_cities, others, key)
        Report.generate_excel(report)
        Report.generate_graph(report)
        Report.generate_pdf(report)


class Report:
    """Класс отвечает за формирование графиков и отчетов

        Attributes:
            salary_filter (dict): Динамика уровня зарплат по годам
            vac_filter (dict): Динамика количества вакансий по годам
            vac_sal_filter (dict): Динамика уровня зарплат по годам для выбранной профессии
            vac_count_filter (dict): Динамика количества вакансий по годам для выбранной профессии
            salary_cities_filter (dict): Уровень зарплат по городам (в порядке убывания)
            vacs_cities (dict): Доля вакансий по городам (в порядке убывания)
            others (float): Доля вакансий по городам не входящих в Топ-10
            vacancy (str): Название профессии
    """
    def __init__(self, salary_filter, vac_filter, vac_sal_filter, vac_count_filter, salary_cities_filter, vacs_cities,
                 others, vacancy):
        """Инициализирует объект Report.

            Args:
                salary_filter (dict): Динамика уровня зарплат по годам
                vac_filter (dict): Динамика количества вакансий по годам
                vac_sal_filter (dict): Динамика уровня зарплат по годам для выбранной профессии
                vac_count_filter (dict): Динамика количества вакансий по годам для выбранной профессии
                salary_cities_filter (dict): Уровень зарплат по городам (в порядке убывания)
                vacs_cities (dict): Доля вакансий по городам (в порядке убывания)
                others (float): Доля вакансий по городам не входящих в Топ-10
                vacancy (str): Название профессии
        """
        self.salary_filter = salary_filter
        self.vac_filter = vac_filter
        self.vac_sal_filter = vac_sal_filter
        self.vac_count_filter = vac_count_filter
        self.salary_cities_filter = salary_cities_filter
        self.vacs_cities = vacs_cities
        self.others = others
        self.vacancy = vacancy

    @staticmethod
    def as_text(value):
        """Парсит value в строку.

            Args:
                value (any): Входное значение, которое нужно конвертировать

            Returns:
                str: value конвертированное в строку.
        """
        if value is None:
            return ''
        return str(value)

    @staticmethod
    def generate_excel(report):
        """Генерирует excel файл с вакансиями

            Args:
                report (Report): Объект класса Report
        """
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'Статистика по годам'
        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {report.vacancy}', 'Количество вакансий',
                  f'Количество вакансий - {report.vacancy}']
        sheet2 = wb.create_sheet('Статистика по городам')
        heads2 = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']

        for i, head in enumerate(heads1):
            sheet1.cell(row=1, column=(i + 1), value=head).font = Font(bold=True)

        for year, value in report.salary_filter.items():
            sheet1.append(
                [year, value, report.vac_sal_filter[year], report.vac_filter[year], report.vac_count_filter[year]])

        for column_cells in sheet1.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column_cells)
            sheet1.column_dimensions[column_cells[0].column_letter].width = length + 2

        thin = Side(border_style='thin', color='000000')
        for column in sheet1.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        for i, head in enumerate(heads2):
            if i + 1 != 3:
                sheet2.cell(row=1, column=(i + 1), value=head).font = Font(bold=True)

        sal_count = len(report.salary_cities_filter)
        if sal_count > 10:
            sal_count = 10

        vac_count = len(report.vacs_cities)
        if vac_count > 10:
            vac_count = 10

        dic = []
        for i, head in enumerate(report.salary_cities_filter.items()):
            if i < sal_count:
                dic.append(list(head) + [''])
            else:
                break

        for i, head in enumerate(report.vacs_cities.items()):
            if i < vac_count:
                dic[i] += list(head)

        for i, value in enumerate(dic):
            sheet2.append(value)
            sheet2[f'E{i + 2}'].number_format = FORMAT_PERCENTAGE_00

        for column_cells in sheet2.columns:
            lenght = max(len(Report.as_text(cell.value)) for cell in column_cells)
            sheet2.column_dimensions[column_cells[0].column_letter].width = lenght + 2

        for row in sheet2.rows:
            for cell in row:
                if cell.value != None and cell.value != '':
                    cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        wb.save('report.xlsx')

    @staticmethod
    def generate_graph(report):
        """Генерирует png файл со статистикой вакансий на графиках

            Args:
                report (Report): Объект класса Report
        """
        width = 0.4
        x_nums = np.arange(len(report.salary_filter.keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2
        fig = plt.figure(figsize=(8, 6))

        ax = fig.add_subplot(221)
        ax.set_title('Уровень зарплат по годам')
        ax.bar(x_list1, report.salary_filter.values(), width, label='средняя з/п')
        ax.bar(x_list2, report.vac_sal_filter.values(), width, label=f'з/п {report.vacancy.lower()}')
        ax.set_xticks(x_nums, report.salary_filter.keys(), rotation='vertical')
        ax.legend(fontsize=8, loc='upper left')
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(222)
        ax.set_title('Количество вакансий по годам')
        ax.bar(x_list1, report.vac_filter.values(), width, label='Количество вакансий')
        ax.bar(x_list2, report.vac_count_filter.values(), width, label=f'Количество вакансий\n{report.vacancy.lower()}')
        ax.set_xticks(x_nums, report.vac_filter.keys(), rotation='vertical')
        ax.legend(fontsize=8, loc='upper left')
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        y = list(reversed(report.salary_cities_filter.keys()))
        y = [x.replace(' ', '\n').replace('-', '-\n') for x in y]
        x = list(reversed(report.salary_cities_filter.values()))
        plt.barh(y, x)
        ax.tick_params(axis='y', labelsize=6)
        ax.tick_params(axis='x', labelsize=8)
        ax.grid(True, axis='x')

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансий по городам")
        city_list = list(report.vacs_cities.keys())
        percent_list = list(report.vacs_cities.values())
        if report.others != 0:
            city_list.insert(0, 'Другие')
            percent_list.insert(0, 1 - sum(percent_list))
        plt.pie(percent_list, labels=city_list, textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')
        # plt.show()

    @staticmethod
    def generate_pdf(report):
        """Генерирует pdf файл из png и excel файлов

            Args:
                report (Report): Объект класса Report
        """
        vacancy = report.vacancy
        image_file = 'graph.png'

        options = {
            "enable-local-file-access": None
        }

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {report.vacancy}', 'Количество вакансий',
                  f'Количество вакансий - {report.vacancy}']
        heads2 = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']
        report.vacs_cities = {key: (str(round(float(value) * 100, 3))).replace('.',',') + '%' for key, value in report.vacs_cities.items()}

        pdf_template = template.render({'vacancy': vacancy, 'image_file': image_file,
                                        "salary_filter": report.salary_filter,
                                        "vac_filter": report.vac_filter,
                                        "vac_sal_filter": report.vac_sal_filter,
                                        "vac_count_filter": report.vac_count_filter,
                                        "salary_cities_filter": report.salary_cities_filter,
                                        "vacs_cities": report.vacs_cities,
                                        "heads1": heads1,
                                        "heads2": heads2})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)


def get_table():
    """Используется в main.py. Формирует pdf файл.
    """
    pars = InputParam()
    if pars.params is not None:
        dataset = DataSet(pars.params[0])
        InputParam.print_data(dataset.vacancies_objects, pars.params[1])
